import pandas as pd
from typing import List, Dict
from pathlib import Path
import io
from datetime import datetime

class OutputGenerator:
    """Generate output files in various formats for RAG pipeline results"""
    
    def __init__(self):
        self.output_dir = Path("output")
        self.output_dir.mkdir(exist_ok=True)
    
    def generate_excel(self, results: List[Dict], filename: str = None) -> str:
        """Generate Excel file with requirements and responses"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"rfp_responses_{timestamp}.xlsx"
        
        # Prepare data for DataFrame
        data = []
        for i, result in enumerate(results, 1):
            row_data = {
                "ID": i,
                "Requirement": result["requirement"],
                "Response": result["response"],
                "Status": result.get("status", "success")
            }
            
            # Add quality scores if available
            if result.get("quality_score") is not None:
                row_data.update({
                    "Quality Score": result["quality_score"],
                    "Quality Status": result.get("quality_status", "Unknown"),
                    "Completeness": result.get("quality_breakdown", {}).get("completeness", ""),
                    "Clarity": result.get("quality_breakdown", {}).get("clarity", ""),
                    "Professionalism": result.get("quality_breakdown", {}).get("professionalism", ""),
                    "Relevance": result.get("quality_breakdown", {}).get("relevance", ""),
                    "Quality Feedback": "; ".join(result.get("quality_feedback", []))
                })
            
            data.append(row_data)
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Save to Excel file
        output_path = self.output_dir / filename
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='RFP Responses', index=False)
            
            # Get the workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['RFP Responses']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set reasonable column widths
                if column_letter == 'A':  # ID column
                    adjusted_width = 5
                elif column_letter == 'B':  # Requirement column
                    adjusted_width = min(max_length + 2, 80)
                elif column_letter == 'C':  # Response column
                    adjusted_width = min(max_length + 2, 100)
                else:  # Status column
                    adjusted_width = min(max_length + 2, 15)
                
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Enable text wrapping for requirement and response columns
            from openpyxl.styles import Alignment
            wrap_alignment = Alignment(wrap_text=True, vertical='top')
            
            for row in worksheet.iter_rows(min_row=2):  # Skip header row
                row[1].alignment = wrap_alignment  # Requirement column
                row[2].alignment = wrap_alignment  # Response column
        
        return str(output_path)
    
    def generate_excel_bytes(self, results: List[Dict]) -> bytes:
        """Generate Excel file as bytes for Streamlit download"""
        # Prepare data for DataFrame
        data = []
        for i, result in enumerate(results, 1):
            row_data = {
                "ID": i,
                "Requirement": result["requirement"],
                "Response": result["response"],
                "Status": result.get("status", "success")
            }
            
            # Add quality scores if available
            if result.get("quality_score") is not None:
                row_data.update({
                    "Quality Score": result["quality_score"],
                    "Quality Status": result.get("quality_status", "Unknown"),
                    "Completeness": result.get("quality_breakdown", {}).get("completeness", ""),
                    "Clarity": result.get("quality_breakdown", {}).get("clarity", ""),
                    "Professionalism": result.get("quality_breakdown", {}).get("professionalism", ""),
                    "Relevance": result.get("quality_breakdown", {}).get("relevance", ""),
                    "Quality Feedback": "; ".join(result.get("quality_feedback", []))
                })
            
            data.append(row_data)
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='RFP Responses', index=False)
            
            # Get the workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['RFP Responses']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set reasonable column widths
                if column_letter == 'A':  # ID column
                    adjusted_width = 5
                elif column_letter == 'B':  # Requirement column
                    adjusted_width = min(max_length + 2, 80)
                elif column_letter == 'C':  # Response column
                    adjusted_width = min(max_length + 2, 100)
                else:  # Status column
                    adjusted_width = min(max_length + 2, 15)
                
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Enable text wrapping for requirement and response columns
            from openpyxl.styles import Alignment
            wrap_alignment = Alignment(wrap_text=True, vertical='top')
            
            for row in worksheet.iter_rows(min_row=2):  # Skip header row
                row[1].alignment = wrap_alignment  # Requirement column
                row[2].alignment = wrap_alignment  # Response column
        
        output.seek(0)
        return output.getvalue()
    
    def generate_csv(self, results: List[Dict], filename: str = None) -> str:
        """Generate CSV file with requirements and responses"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"rfp_responses_{timestamp}.csv"
        
        # Prepare data for DataFrame
        data = []
        for i, result in enumerate(results, 1):
            data.append({
                "ID": i,
                "Requirement": result["requirement"],
                "Response": result["response"],
                "Status": result.get("status", "success")
            })
        
        # Create DataFrame and save as CSV
        df = pd.DataFrame(data)
        output_path = self.output_dir / filename
        df.to_csv(output_path, index=False)
        
        return str(output_path)
    
    def generate_csv_bytes(self, results: List[Dict]) -> bytes:
        """Generate CSV file as bytes for Streamlit download"""
        # Prepare data for DataFrame
        data = []
        for i, result in enumerate(results, 1):
            data.append({
                "ID": i,
                "Requirement": result["requirement"],
                "Response": result["response"],
                "Status": result.get("status", "success")
            })
        
        # Create DataFrame and convert to CSV bytes
        df = pd.DataFrame(data)
    def generate_structured_excel_bytes(self, results: List[Dict], original_df: pd.DataFrame, 
                                       requirement_column: str) -> bytes:
        """Generate Excel file preserving original structure with added response column"""
        # Create a copy of the original dataframe
        output_df = original_df.copy()
        
        # Create a mapping of requirements to responses
        response_map = {result["requirement"]: result["response"] for result in results}
        
        # Add response column
        responses = []
        for _, row in output_df.iterrows():
            requirement = str(row[requirement_column]).strip()
            response = response_map.get(requirement, "No response generated")
            responses.append(response)
        
        output_df['Response'] = responses
        
        # Add status column
        statuses = []
        for _, row in output_df.iterrows():
            requirement = str(row[requirement_column]).strip()
            status = next((result["status"] for result in results 
                          if result["requirement"] == requirement), "unknown")
            statuses.append(status)
        
        output_df['Status'] = statuses
        
        # Create Excel file in memory
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            output_df.to_excel(writer, sheet_name='RFP Responses', index=False)
            
            # Get the workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['RFP Responses']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set reasonable column widths
                if column_letter in ['A', 'B', 'C']:  # First few columns
                    adjusted_width = min(max_length + 2, 30)
                else:  # Response and status columns
                    adjusted_width = min(max_length + 2, 100)
                
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Enable text wrapping for response column
            from openpyxl.styles import Alignment
            wrap_alignment = Alignment(wrap_text=True, vertical='top')
            
            # Find response column (should be second to last)
            response_col_idx = len(output_df.columns) - 1  # -1 for Response, -2 would be Status
            
            for row in worksheet.iter_rows(min_row=2):  # Skip header row
                if len(row) > response_col_idx:
                    row[response_col_idx - 1].alignment = wrap_alignment  # Response column
        
        output.seek(0)
        return output.getvalue()